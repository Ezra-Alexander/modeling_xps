import numpy as np
import sys
import math
from geom_helper import *
from openpyxl import Workbook, load_workbook
from ddft_helper import get_energy
from qchem_helper import read_lowdin, get_geom_io

# the main juice of the bash script of the same name
# for a given ddft excited state, this script:
#	1. Computes the CEBE
#   2. Confirms that the correct P was excited
#		2.a. If wrong P was excited or CEBE is too high, prints stuff and skips excel editing
#   3. Writes the results to an excel script
#   	3.a. Creates the excel script if necessary
#   4. weights the different peaks by the appropriate amount using the directory name and a read-in weight
#		4.a. This means it extracts the P info from the geometry
#		4.b. Labels each peak as bulk, P-3c, or describing surface In
#		4.c. I'm assuming that P-3c and surface P-4c should be weighted the same but that isn't necessarily correct
#		4.d. I might also want to change "surface P-4c" to just "P-4c on In-3c and/or P-4c on high ligand In"
# right now this is hard-coded for P but it wouldn't be hard to change

es_out=sys.argv[1] #excited state qchem ddft.out
gs_out=sys.argv[2] #excited state qchem ddft.out
iteration=int(sys.argv[3]) #which step of the excel creation you are on. For indexing
directory=sys.argv[4] #name of the directory, which should be of the form P_{N}, where N is the P-specific index.
ligand=sys.argv[5] #the ligand used here. Abbreviate. "F" for fluorine, "Cl" for chlorine
bulk_weight=int(sys.argv[6]) #the weighting to be applied to bulk P relative to P-uc and surface P-4c

energy_cutoff=139 #empirical parameter for upper bound on expected CEBEs. CEBEs above this will be considered to have failed to converge
bond_cutoff=2.75 #empirical parameter for upper bound on chemical bonds

#compute the CEBE
gs_energy = get_energy(gs_out) 
es_energy = get_energy(es_out)
energy_difference = es_energy - gs_energy

#determine target P
p_index=directory[2:-1] #note - leaving as a string

#get geometry
atoms,coords=get_geom_io(gs_out)

#read Lowdin populations
#read first orb on the target
orb_num=read_lowdin(es_out,"P",p_index,2,"p") 
#compare to minimum ground state P 2p orb
min_p2p=read_lowdin(gs_out,"P",str(0),2,"p")
if orb_num<min_p2p and energy_difference<energy_cutoff: #1
	print("Excitation of", directory[:-1], "successful with energy", energy_difference, "and excited index",orb_num)
	converged=True
else:
	print("Warning! Excitation of", directory[:-1], "unsuccessful with energy", energy_difference, "and excited index",orb_num)
	converged=False

if iteration==0:
	wb=Workbook()
	ws1 = wb.active #The 1st sheet contains all the excitations
	ws1.title="Full"
	ws2 = wb.create_sheet("Sorted")
	#write column labels
	ws2["A1"]="Hole Traps"
	ws2["B1"]="Bulk Phosphorus"	
	ws2["C1"]="Phosphorus Bound to Electron Traps"	
	ws2["D1"]="Phosphorus Bound to Dopant"	
	ws1["A1"]="Excitation"
	ws1["B1"]="Atom"
	ws1["C1"]="Number"
	ws1["D1"]="CEBE (eV)"
	ws1["E1"]="Location"
	#not going to write timing data but it wouldn't be hard to add
else:
	wb = load_workbook("collected_spectrum.xlsx")
	ws1 = wb["Full"] #The 1st sheet contains all the excitations
	ws2=wb["Sorted"]

if converged==True:

	#determine what type of P the target is
	label=get_label(atoms,coords,"P",int(p_index),bond_cutoff,ligand)
	print(label)

	#now we need to write this to the excel script
	#let's try to stick with the format I used before
	#which means we need two sheets
	alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
	
	#start with sheet 1, which contains all excitations
	#determine the first blank row
	active_row=1
	for row in ws1.values:
		active_row=active_row+1

	#now write the data
	if label=="Bulk":
		i=0
		while i<bulk_weight:
			ws1["A"+str(active_row)]=orb_num
			ws1["B"+str(active_row)]="P"
			ws1["C"+str(active_row)]=p_index
			ws1["D"+str(active_row)]=energy_difference
			ws1["E"+str(active_row)]=label
			active_row=active_row+1
			i=i+1
	else:
		ws1["A"+str(active_row)]=orb_num
		ws1["B"+str(active_row)]="P"
		ws1["C"+str(active_row)]=p_index
		ws1["D"+str(active_row)]=energy_difference
		ws1["E"+str(active_row)]=label
		active_row=active_row+1

	#now do the second sheet, where everything is sorted into columns

	if label=="Bulk":
		column="B"
	elif label=="Under-Coordinated":
		column="A"
	elif label=="Bound to Trap":
		column="C"
	elif label=="Bound to Dopant":
		column="D"
	else:
		print("Something went wrong with labeling")

	active_row=1
	for row in ws2.values:
		if ws2[column+str(active_row)].value:
			active_row=active_row+1
			
	#write data
	if label=="Bulk":
		i=0
		while i<bulk_weight:
			ws2[column+str(active_row)]=energy_difference
			active_row=active_row+1
			i=i+1
	else:
		ws2[column+str(active_row)]=energy_difference
		active_row=active_row+1


	wb.save("collected_spectrum.xlsx")


else: #I could write a new input that uses different convergence paramters?
	print("A new input file has been written for the excitation of", directory[:-1]) #note. I have not actually done this yet

