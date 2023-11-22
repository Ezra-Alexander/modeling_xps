import numpy as np
from pdos_helper import dos_grid_general,get_alpha,get_ao_ind, get_ind_ao_underc
from matplotlib import pyplot as plt
import random
import pandas as pd
import sys
import math
import scipy.special
from openpyxl import Workbook, load_workbook
from geom_helper import *

#this script should take an excel file with all the necessary XPS data (that I have already made) and apply a given, fixed spin-orbit splitting
#What I want to do is average the CEBEs for each P, then apply the 2:1 splitting to each
#then write this new data to a new, similar excel titled "spin_orbit_[old_excel_name.xlsx]"
#honestly pretty simple
#note that this is the "easy" (i.e. lazy) way to do S.O. coupling

excel=sys.argv[1] #the excel file with the XPS data in it. Should be formatted with the first page containing all excitations (Excitation, Atom, Number, CEBE, Time, Test) and the second page containing the relevant subcategories. May be a third page, but this won't read that
splitting=float(sys.argv[2]) #the given spin-orbit P 2p splitting to apply. Taken from experiment. Using 0.86 eV from PH3[1] & PF3[2]
#[1] Cavell, R. G.; Tan, K. H. Chemical Physics Letters 1992.
#[2] Väyrynen, I. J et al. J. Electron Spectrosc. Relat. Phenom. 1992.

#read the 1st sheet, with the full XPS
full=pd.read_excel(excel,sheet_name=0)
p_indices_col=full.columns[2]
p_indices=[]
for i, row in full.iterrows():
	p_indices.append(int(row[p_indices_col]))
cebes_col=full.columns[3]
cebes=[]
for i, row in full.iterrows():
	cebes.append(float(row[cebes_col]))
n_phos=max(p_indices)

# #go through and group the CEBEs by P
# cebes_by_p=[[]for x in range(n_phos)]
# for i,cebe in enumerate(cebes):
# 	cebes_by_p[p_indices[i]-1].append(cebe)

# #average cebes
# averages=[]
# for i,p in enumerate(cebes_by_p):
# 	if len(p)>1:
# 		averages.append(sum(p)/len(p))
# 	elif len(p)==1:
# 		averages.append(p[0])

#split 'em
split_cebes=[]
for i,av in enumerate(cebes):
	p232=av-(splitting/3)
	p212=av+(2*splitting/3)
	split_cebes.append(p232)
	split_cebes.append(p232)
	split_cebes.append(p212)

#write this to the new excel file
excel_name="spin_orbit_"+excel
wb=Workbook()
ws = wb.active
ws['A1']="Excitation"	
ws['B1']='Atom'
ws['C1']='Number'
ws['D1']='CEBE (eV)'
ws['E1']='Time'
ws['F1']='Test'
for i,cebe in enumerate(split_cebes):
	ws['D'+str(i+2)]=cebe

#Now we do the individual sub-peaks
ws2=wb.create_sheet("Assignments")

#read sheet 2 of original excel
sub = pd.read_excel(excel,sheet_name=1)
sub_cebes = [[] for i in range(sub.shape[1])]
i = 0
names = []
for name,data in sub.iteritems():
	names.append(name)
	for cebe in data.values:
		if not math.isnan(cebe):
			sub_cebes[i].append(float(cebe))
	i = i+1


#make said subcategories in the same order as in the new excel, but first the 1/2 and then the 3/2
alphabet="ABCDEFGHIJKLMNOPQRSTUVWXYZ" #lol
for i,name in enumerate(names):
	ws2[alphabet[i]+str(1)]=name+" 2p 3/2"
	ws2[alphabet[i+len(names)]+str(1)]=name+" 2p 1/2"

#now what we want to do is take each subcategory, determine the P in it, and assign the new values to that subcategory in the new excel
for i,name in enumerate(names):
	p_in_sub=[]
	for j,cebe in enumerate(sub_cebes[i]):
		#determine which P the cebe belongs to
		p_num=p_indices[cebes.index(cebe)]
		if p_num not in p_in_sub:
			p_in_sub.append(p_num)

	row=2

	for k,new_cebe in enumerate(split_cebes):
			if k%3 ==2:
				if (p_indices[cebes.index(new_cebe-(2*splitting/3))] in p_in_sub):
					ws2[alphabet[i+len(names)]+str(row)]=new_cebe
			else:
				if (p_indices[cebes.index(new_cebe+(splitting/3))] in p_in_sub):
					ws2[alphabet[i]+str(row)]=new_cebe
			row = row+1


wb.save(excel_name)
